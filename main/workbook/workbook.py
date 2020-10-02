from openpyxl import Workbook
from cloudflare.cloudflare import Zone
from cloudflare.cloudflare import Record
import cloudflare.cloudflare as cloudflare


def workbook() -> Workbook:
    wbook = Workbook()
    wbook.remove(wbook.active)

    # Create Zones worksheet
    wsheet = wbook.create_sheet("Zones")
    wsheet.append(cloudflare.zones_header)

    return wbook


def zone(wbook: Workbook, value: Zone) -> None:
    wbook["Zones"].append(value.row())


def record(wbook: Workbook, parent: Zone, value: Record) -> None:
    wsheet = None
    wsheet_name = parent.name[:31]

    # Find worksheet with Zone name
    for ws in wbook.worksheets:
        if ws.title == wsheet_name:
            wsheet = ws
            break

    # Create worksheet when not found
    if wsheet is None:
        wsheet = wbook.create_sheet(wsheet_name)
        wsheet.append(cloudflare.records_header)

    # Append data
    wsheet.append(value.row())


def save(wbook: Workbook, filename: str) -> bool:
    try:
        return wbook.save(filename)
    except TypeError:
        return False
